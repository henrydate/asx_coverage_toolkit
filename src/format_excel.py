"""
src/format_excel.py
────────────────────
Builds the professionally formatted ASX Master Database Excel workbook.
Takes the enriched DataFrame from enrich.py and produces a 3-sheet XLSX:
    1. ASX Database    — main data, colour-coded, auto-filter, dropdowns
    2. Summary         — pivot counts by sector, entity type, market cap, index
    3. Legend          — column guide and data quality disclaimer
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Colour palette ────────────────────────────────────────────────────────────
NAVY   = "1B3A6B"
WHITE  = "FFFFFF"
GOLD   = "C9A84C"
DGREY  = "2D2D2D"

# Sector row colours: (light_bg_hex, text_hex)
SEC_CLR: dict[str, tuple[str, str]] = {
    "Financials":              ("D4E6F7", "1B5E96"),
    "Materials":               ("F5EAD8", "6D3813"),
    "Energy":                  ("FFF2D9", "CC5500"),
    "Information Technology":  ("E8EDFB", "1A237E"),
    "Health Care":             ("E4F5EC", "145A32"),
    "Industrials":             ("ECEFF1", "455A64"),
    "Consumer Discretionary":  ("F3E5F5", "6A1B9A"),
    "Consumer Staples":        ("E0F4E8", "1B5E20"),
    "Communication Services":  ("E0F7FA", "006064"),
    "Real Estate":             ("FBF3ED", "4E342E"),
    "Utilities":               ("ECEFF1", "37474F"),
    "Unclassified":            ("F5F5F5", "757575"),
}

CAP_CLR: dict[str, str] = {
    "Mega Cap":  "1B3A6B",
    "Large Cap": "1565C0",
    "Mid Cap":   "0288D1",
    "Small Cap": "00897B",
    "Micro Cap": "558B2F",
    "Nano Cap":  "F57F17",
    "Unknown":   "9E9E9E",
}

BHS_FILL:  dict[str, str] = {"Buy": "C8E6C9", "Hold": "FFF9C4", "Sell": "FFCDD2"}
PROG_FILL: dict[str, str] = {"TA": "D1ECF1", "FA": "D4EDDA", "TA + FA": "FFF3CD"}
SRC_FILL:  dict[str, str] = {
    "Research Data":   "E8F5E9",
    "Live Data":       "E3F2FD",
    "Source CSV":      "E3F2FD",
    "Sector Estimate": "FFF8E1",
}

# ── Column definitions: (name, width) ────────────────────────────────────────
COLUMNS: list[tuple[str, int]] = [
    ("Company",                       42),
    ("ASX Ticker",                    10),
    ("Entity Type",                   22),
    ("Market Cap Category",           16),
    ("GICS Sector",                   24),
    ("GICS Industry Group",           28),
    ("GICS Industry",                 34),
    ("GICS Sub-Industry",             34),
    ("Market Cap (AUD)",              16),
    ("Market Cap Data Date",          18),
    ("Share Price (AUD)",             14),
    ("Shares Outstanding",            16),
    ("Revenue (AUD, Latest Annual)",  22),
    ("Net Profit/Loss (AUD)",         18),
    ("Dividend Yield",                12),
    ("Franking Level",                12),
    ("P/E Ratio",                     10),
    ("Price/Book",                    10),
    ("Beta (vs ASX 200)",             13),
    ("Return on Equity (ROE)",        15),
    ("Debt/Equity Ratio",             13),
    ("Reporting Currency",            16),
    ("Fiscal Year End",               13),
    ("CEO / MD",                      28),
    ("Website",                       28),
    ("Year Founded",                  12),
    ("Business Description",          55),
    ("HQ State / Territory",          16),
    ("Primary Country of Operations", 24),
    ("Domicile",                      16),
    ("Primary Exchange",              14),
    ("Dual Listed",                   10),
    ("Dual Listing Exchange",         16),
    ("Listing Status",                13),
    ("ASX 20",                         8),
    ("ASX 50",                         8),
    ("ASX 100",                        9),
    ("ASX 200",                        9),
    ("All Ordinaries",                14),
    ("Index Membership",              38),
    ("Data Source",                   16),
    ("BHS Rating",                    12),
    ("Progress",                      14),
    ("Notes",                         55),
]

COL_NAMES = [c[0] for c in COLUMNS]

RIGHT_COLS = {
    "Dividend Yield", "Franking Level", "P/E Ratio", "Price/Book",
    "Beta (vs ASX 200)", "Return on Equity (ROE)", "Debt/Equity Ratio",
    "Market Cap (AUD)", "Revenue (AUD, Latest Annual)", "Net Profit/Loss (AUD)",
    "Share Price (AUD)", "Shares Outstanding",
}
CENTER_COLS = {
    "ASX Ticker", "ASX 20", "ASX 50", "ASX 100", "ASX 200", "All Ordinaries",
    "Dual Listed", "BHS Rating", "Progress", "Data Source", "Market Cap Category",
    "Listing Status", "Reporting Currency", "Fiscal Year End",
    "Dual Listing Exchange", "Primary Exchange",
}


# ── Style helpers ─────────────────────────────────────────────────────────────

def _hf(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _thin_border() -> Border:
    t = Side(style="thin", color="DDDDDD")
    return Border(left=t, right=t, top=t, bottom=t)

def _header_border() -> Border:
    t = Side(style="thin", color="DDDDDD")
    b = Side(style="medium", color=GOLD)
    return Border(left=t, right=t, top=t, bottom=b)


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_database_sheet(ws, df: pd.DataFrame, as_of: str) -> None:
    """Build the main ASX Database sheet."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Title row
    n_cols = len(COLUMNS)
    ws.row_dimensions[1].height = 32
    tc = ws.cell(1, 1, f"ASX MASTER DATABASE — {len(df):,} Entities | {n_cols} Columns | as at {as_of}")
    tc.font = Font(name="Arial", bold=True, size=13, color=WHITE)
    tc.fill = _hf(NAVY)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    # Header row
    ws.row_dimensions[2].height = 38
    hdr_border = _header_border()
    for ci, (cn, cw) in enumerate(COLUMNS, 1):
        c = ws.cell(2, ci, cn)
        c.font = Font(name="Arial", bold=True, size=8.5, color=WHITE)
        c.fill = _hf(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = hdr_border
        ws.column_dimensions[get_column_letter(ci)].width = cw

    # Dropdown validation
    bhs_col  = COL_NAMES.index("BHS Rating") + 1
    prog_col = COL_NAMES.index("Progress") + 1
    last_row = len(df) + 2

    bhs_dv = DataValidation(
        type="list", formula1='"Buy,Hold,Sell"', allow_blank=True, showDropDown=False,
        prompt="Select analyst rating", promptTitle="BHS Rating",
    )
    prog_dv = DataValidation(
        type="list", formula1='"TA,FA,TA + FA"', allow_blank=True, showDropDown=False,
        prompt="TA=Technical done; FA=Fundamental done", promptTitle="Analysis Progress",
    )
    ws.add_data_validation(bhs_dv)
    ws.add_data_validation(prog_dv)
    bhs_dv.sqref  = f"{get_column_letter(bhs_col)}3:{get_column_letter(bhs_col)}{last_row}"
    prog_dv.sqref = f"{get_column_letter(prog_col)}3:{get_column_letter(prog_col)}{last_row}"

    std_b = _thin_border()

    # Data rows
    for ri, (_, row) in enumerate(df.iterrows(), 3):
        sector  = str(row.get("GICS Sector", "Unclassified"))
        bg_hex, _ = SEC_CLR.get(sector, ("FFFFFF", "000000"))
        alt_bg  = "FFFFFF" if ri % 2 == 0 else bg_hex
        ws.row_dimensions[ri].height = 15

        for ci, (cn, _) in enumerate(COLUMNS, 1):
            val = str(row.get(cn, ""))
            if val in ("nan", "None"):
                val = ""
            c = ws.cell(ri, ci, val)
            c.border = std_b

            # Defaults
            fill_hex = alt_bg
            fnt_color, fnt_bold, fnt_size, fnt_italic = DGREY, False, 8.5, False
            h_align = "right" if cn in RIGHT_COLS else ("center" if cn in CENTER_COLS else "left")

            # Column-specific overrides
            if cn == "ASX Ticker":
                fnt_color, fnt_bold = NAVY, True

            elif cn == "Market Cap Category":
                clr = CAP_CLR.get(val, "9E9E9E")
                c.font = Font(name="Arial", size=8, bold=True, color=WHITE)
                c.fill = _hf(clr)
                c.alignment = Alignment(horizontal="center", vertical="center")
                continue

            elif cn == "GICS Sector" and sector != "Unclassified":
                _, txt = SEC_CLR.get(sector, ("FFFFFF", "000000"))
                fnt_color, fnt_bold = txt, True

            elif cn == "BHS Rating":
                bhs_c = BHS_FILL.get(val)
                if bhs_c:
                    fill_hex = bhs_c
                fnt_bold = True

            elif cn == "Progress":
                pc = PROG_FILL.get(val)
                if pc:
                    fill_hex = pc
                fnt_bold, fnt_color = True, "006B3C"

            elif cn in ("ASX 20", "ASX 50", "ASX 100", "ASX 200", "All Ordinaries"):
                if val == "Yes":
                    fill_hex, fnt_bold, fnt_color = "C8E6C9", True, "1B5E20"
                else:
                    fill_hex, fnt_color, fnt_size = "FAFAFA", "CCCCCC", 8.0

            elif cn == "Data Source":
                sc = SRC_FILL.get(val, "FFFFFF")
                fill_hex = sc
                if val == "Sector Estimate":
                    fnt_italic, fnt_color, fnt_size = True, "888888", 8.0
                elif val == "Research Data":
                    fnt_bold, fnt_color, fnt_size = True, "1B5E20", 8.0
                else:
                    fnt_color, fnt_size = "1565C0", 8.0

            elif cn == "Dual Listed" and val == "Yes":
                fill_hex, fnt_color, fnt_bold = "FFF3E0", "E65100", True

            elif cn == "Business Description":
                fnt_italic, fnt_color, fnt_size = True, "555555", 8.0

            elif cn == "Notes":
                fnt_size = 8.0
                if val == "No flags":
                    fnt_color = "BBBBBB"
                elif "Sector Estimate" in val:
                    fnt_italic, fnt_color = True, "999999"
                elif "pending" in val.lower():
                    fnt_color = "CC7000"
                else:
                    fnt_color = "666666"

            c.font = Font(name="Arial", size=fnt_size, bold=fnt_bold,
                          italic=fnt_italic, color=fnt_color)
            c.fill = _hf(fill_hex)
            c.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=False)

    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}{last_row}"


def _build_summary_sheet(ws, df: pd.DataFrame, as_of: str) -> None:
    """Build the Summary dashboard sheet."""
    ws.sheet_view.showGridLines = False

    for col, w in [("A", 30), ("B", 15), ("C", 3), ("D", 30), ("E", 15), ("F", 3), ("G", 30), ("H", 15)]:
        ws.column_dimensions[col].width = w

    std_b = _thin_border()

    def sh(r: int, c: int, txt: str, span: int = 1, bg: str = NAVY) -> None:
        cell = ws.cell(r, c, txt)
        cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        cell.fill = _hf(bg)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 22
        if span > 1:
            ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + span - 1)

    def sr(r: int, c: int, lbl: str, val: int | str, bg: str = "FFFFFF") -> None:
        lc = ws.cell(r, c, lbl)
        vc = ws.cell(r, c + 1, val)
        lc.font = Font(name="Arial", size=9, color=DGREY)
        lc.fill = _hf(bg); lc.border = std_b
        lc.alignment = Alignment(indent=1, vertical="center")
        vc.font = Font(name="Arial", bold=True, size=9, color=NAVY)
        vc.fill = _hf(bg); vc.border = std_b
        vc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.row_dimensions[r].height = 17

    r = 1
    ws.cell(r, 1, "ASX MASTER DATABASE — SUMMARY").font = Font(
        name="Arial", bold=True, size=14, color=NAVY
    )
    ws.merge_cells(f"A{r}:H{r}")
    ws.row_dimensions[r].height = 30
    r += 1
    ws.cell(r, 1, f"As at {as_of}  |  {len(df):,} entities  |  {len(df.columns)} columns").font = Font(
        name="Arial", italic=True, size=9, color="666666"
    )
    r += 2

    # Left: sector
    sh(r, 1, "BY GICS SECTOR", 2); r += 1
    for sec, cnt in df["GICS Sector"].value_counts().items():
        bg, txt = SEC_CLR.get(str(sec), ("FFFFFF", "000000"))
        lc = ws.cell(r, 1, str(sec))
        vc = ws.cell(r, 2, int(cnt))
        for cx in (lc, vc):
            cx.fill = _hf(bg); cx.border = std_b
            cx.alignment = Alignment(indent=1, vertical="center")
        lc.font = Font(name="Arial", bold=True, size=9, color=txt)
        vc.font = Font(name="Arial", bold=True, size=9, color=NAVY)
        vc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.row_dimensions[r].height = 16; r += 1

    r += 1; sh(r, 1, "BY ENTITY TYPE", 2); r += 1
    for et, cnt in df["Entity Type"].value_counts().items():
        sr(r, 1, str(et), int(cnt), "FAFAFA"); r += 1

    # Middle: cap category
    r2 = 4; sh(r2, 4, "BY MARKET CAP CATEGORY", 2); r2 += 1
    for cap, cnt in df["Market Cap Category"].value_counts().items():
        clr = CAP_CLR.get(str(cap), "9E9E9E")
        lc = ws.cell(r2, 4, str(cap))
        vc = ws.cell(r2, 5, int(cnt))
        lc.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        lc.fill = _hf(clr); lc.border = std_b
        lc.alignment = Alignment(indent=1, vertical="center")
        vc.font = Font(name="Arial", bold=True, size=9, color=NAVY)
        vc.fill = _hf("F8F8F8"); vc.border = std_b
        vc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.row_dimensions[r2].height = 16; r2 += 1

    r2 += 2; sh(r2, 4, "INDEX MEMBERSHIP", 2); r2 += 1
    for idx_col, lbl in [
        ("ASX 20", "S&P/ASX 20"), ("ASX 50", "S&P/ASX 50"), ("ASX 100", "S&P/ASX 100"),
        ("ASX 200", "S&P/ASX 200"), ("All Ordinaries", "All Ordinaries"),
    ]:
        cnt = int(df[df[idx_col] == "Yes"].shape[0])
        sr(r2, 4, lbl, cnt, "E8F4FE"); r2 += 1

    r2 += 2; sh(r2, 4, "DATA SOURCE", 2); r2 += 1
    for src, cnt in df["Data Source"].value_counts().items():
        bg = SRC_FILL.get(str(src), "FFFFFF")
        sr(r2, 4, str(src), int(cnt), bg); r2 += 1

    r2 += 2; sh(r2, 4, "OVERVIEW", 2); r2 += 1
    sr(r2, 4, "Total Entities", len(df), "F0F4FF"); r2 += 1
    sr(r2, 4, "With Research / Live Data",
       int(df[df["Data Source"] != "Sector Estimate"].shape[0]), "E8F5E9"); r2 += 1
    sr(r2, 4, "Market Cap Known",
       int(df[df["Market Cap (AUD)"].str.strip().str.len() > 0].shape[0]), "E8F5E9"); r2 += 1
    sr(r2, 4, "Sector Estimates",
       int(df[df["Data Source"] == "Sector Estimate"].shape[0]), "FFF8E1"); r2 += 1
    sr(r2, 4, "Dual-Listed",
       int(df[df["Dual Listed"] == "Yes"].shape[0]), "FFF3E0"); r2 += 1

    # Right: HQ state
    r3 = 4; sh(r3, 7, "BY HQ STATE / TERRITORY", 2); r3 += 1
    for st, cnt in df["HQ State / Territory"].value_counts().head(15).items():
        if st and st not in ("N/A", ""):
            sr(r3, 7, str(st), int(cnt), "FAFAFA"); r3 += 1


def _build_legend_sheet(ws) -> None:
    """Build the column guide and disclaimer sheet."""
    ws.sheet_view.showGridLines = False
    for col, w in [("A", 30), ("B", 52), ("C", 26)]:
        ws.column_dimensions[col].width = w

    std_b = _thin_border()

    def leg_hdr(r: int, txt: str, bg: str = NAVY) -> None:
        for ci in range(1, 4):
            c = ws.cell(r, ci)
            c.fill = _hf(bg); c.border = std_b
        ws.cell(r, 1, txt).font = Font(name="Arial", bold=True, size=11, color=WHITE)
        ws.cell(r, 1).alignment = Alignment(indent=1, vertical="center")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.row_dimensions[r].height = 22

    def leg_row(r: int, name: str, desc: str, ex: str = "", alt: bool = False) -> None:
        bg = "FAFAFA" if alt else "FFFFFF"
        for ci, val in enumerate([name, desc, ex], 1):
            c = ws.cell(r, ci, val)
            c.font = Font(name="Arial", size=9, color=DGREY, bold=(ci == 1), italic=(ci == 3))
            c.fill = _hf(bg); c.border = std_b
            c.alignment = Alignment(indent=1, vertical="center", wrap_text=(ci == 2))
        ws.row_dimensions[r].height = 18

    r = 1
    ws.cell(r, 1, "ASX DATABASE — COLUMN GUIDE & DATA QUALITY KEY").font = Font(
        name="Arial", bold=True, size=13, color=NAVY
    )
    ws.merge_cells(f"A{r}:C{r}"); ws.row_dimensions[r].height = 30; r += 2

    leg_hdr(r, "DATA SOURCE KEY"); r += 1
    for src_lbl, desc, bg in [
        ("Research Data",   "Curated company profile + research-verified fundamentals; price & market cap refreshed live where available.", "C8E6C9"),
        ("Live Data",       "Fetched live from yfinance this session. Medium confidence — may be stale for small-caps.", "BBDEFB"),
        ("Sector Estimate", "No individual data found. Values are SECTOR-LEVEL DEFAULTS only. Verify before acting.", "FFF9C4"),
    ]:
        for ci, val in enumerate([src_lbl, desc], 1):
            c = ws.cell(r, ci, val)
            c.font = Font(name="Arial", size=9, bold=(ci == 1))
            c.fill = _hf(bg); c.border = std_b
            c.alignment = Alignment(indent=1, vertical="center", wrap_text=(ci == 2))
        ws.row_dimensions[r].height = 26; r += 1

    r += 1; leg_hdr(r, "COLUMN REFERENCE"); r += 1
    sub_bg = "3D5A8A"
    for ci, hdr in enumerate(["Column Name", "Description", "Example / Notes"], 1):
        c = ws.cell(r, ci, hdr)
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = _hf(sub_bg); c.border = std_b
        c.alignment = Alignment(indent=1, vertical="center")
    ws.row_dimensions[r].height = 16; r += 1

    col_defs = [
        ("Company",                      "Full legal entity name (cleaned)",                           "COMMONWEALTH BANK OF AUSTRALIA"),
        ("ASX Ticker",                   "ASX exchange ticker symbol (2–5 chars)",                     "CBA"),
        ("Entity Type",                  "Vehicle type: Company, REIT, ETF, LIC, ABS Trust, etc.",     "REIT"),
        ("Market Cap Category",          "Mega≥$50B | Large $10-50B | Mid $2-10B | Small $0.3-2B | Micro $0.05-0.3B | Nano<$0.05B", "Large Cap"),
        ("GICS Sector",                  "Tier 1 GICS (11 sectors)",                                   "Financials"),
        ("GICS Industry Group",          "Tier 2 GICS (25 groups)",                                    "Banks"),
        ("GICS Industry",                "Tier 3 GICS (74 industries)",                                "Diversified Banks"),
        ("GICS Sub-Industry",            "Tier 4 GICS (163 sub-industries); Multiple = conglomerate",  "Regional Banks"),
        ("Market Cap (AUD)",             "Total market cap in AUD at data date; N/A if unavailable",   "A$289.200B"),
        ("Market Cap Data Date",         "Vintage of the market-cap figure: live fetch date or curated snapshot", "Jun-2026 / Apr-2026"),
        ("Share Price (AUD)",            "Share price in AUD at data date",                            "A$172.8000"),
        ("Shares Outstanding",           "Total shares on issue",                                      "1.673B"),
        ("Revenue (AUD, Latest Annual)", "Most recent full-year revenue; sector default if no data",    "A$26.700B"),
        ("Net Profit/Loss (AUD)",        "NPAT latest full year; negative = loss",                     "A$10.200B"),
        ("Dividend Yield",               "Annual gross dividend / share price (%)",                    "3.8%"),
        ("Franking Level",               "Percentage franking of dividends (0–100%)",                  "100%"),
        ("P/E Ratio",                    "Price-to-earnings ratio (trailing)",                         "22.5x"),
        ("Price/Book",                   "Price-to-book value ratio",                                  "2.8x"),
        ("Beta (vs ASX 200)",            "12-month beta relative to S&P/ASX 200",                      "0.85"),
        ("Return on Equity (ROE)",       "Net profit / avg shareholders equity (%)",                   "14.5%"),
        ("Debt/Equity Ratio",            "Total debt / total equity",                                  "0.15x"),
        ("Reporting Currency",           "Primary currency for financial statements",                  "AUD, USD, NZD"),
        ("Fiscal Year End",              "Month the company's financial year ends",                    "Jun"),
        ("CEO / MD",                     "Current Chief Executive Officer or Managing Director",        "Matt Comyn"),
        ("Website",                      "Investor relations website",                                 "commbank.com.au"),
        ("Year Founded",                 "Year the company was established",                           "1911"),
        ("Business Description",         "One-sentence primary business summary",                      "Australia's largest bank..."),
        ("HQ State / Territory",         "Australian state of head office",                            "NSW, VIC, WA..."),
        ("Primary Country of Operations","Main geographic focus",                                      "Australia, Global..."),
        ("Domicile",                     "Country of incorporation / tax domicile",                    "Australia, USA, NZ..."),
        ("Primary Exchange",             "Exchange where primarily listed",                            "ASX"),
        ("Dual Listed",                  "Secondary listing on another exchange?",                     "Yes / No"),
        ("Dual Listing Exchange",        "Secondary exchange(s) if dual listed",                       "NYSE, LSE, NZX..."),
        ("Listing Status",               "Current ASX status",                                         "Listed"),
        ("ASX 20",                       "Member of S&P/ASX 20?",                                      "Yes / No"),
        ("ASX 50",                       "Member of S&P/ASX 50?",                                      "Yes / No"),
        ("ASX 100",                      "Member of S&P/ASX 100?",                                     "Yes / No"),
        ("ASX 200",                      "Member of S&P/ASX 200? (benchmark index)",                  "Yes / No"),
        ("All Ordinaries",               "Member of the All Ordinaries (~500 largest)?",               "Yes / No"),
        ("Index Membership",             "Concatenated membership string for filtering",               "ASX 20 | ASX 50 | ASX 100"),
        ("Data Source",                  "Research Data=verified | Live Data=yfinance | Sector Estimate=defaults", "Research Data"),
        ("BHS Rating",                   "YOUR analyst rating — dropdown: Buy / Hold / Sell",          "Buy"),
        ("Progress",                     "Analysis stage — TA: Technical done; FA: Fundamental done", "TA + FA"),
        ("Notes",                        "Auto-generated flags: unclassified GICS, estimates, foreign domicile, dual-listed", "No flags"),
    ]

    for i, (name, desc, ex) in enumerate(col_defs):
        leg_row(r, name, desc, ex, alt=(i % 2 == 0)); r += 1

    r += 1
    leg_hdr(r, "⚠️  IMPORTANT DISCLAIMER", "AA3300"); r += 1
    disc = (
        'Rows tagged "Sector Estimate" use ILLUSTRATIVE DEFAULTS based on typical sector medians — '
        "they are NOT company-specific financial data. For ~1,800+ smaller entities (primarily "
        "micro/nano-cap explorers and special vehicles), no public bulk financial source is available. "
        "Revenue, profit, P/E, ROE, Beta and Debt/Equity for these rows are approximations only. "
        "Always verify using ASX announcements, company websites, or a licensed data provider "
        "(Bloomberg, FactSet, Morningstar) before making any investment decisions. "
        "This tool is for research-triage and screening purposes only."
    )
    c = ws.cell(r, 1, disc)
    c.font = Font(name="Arial", size=9, italic=True, color="880000")
    c.fill = _hf("FFF0F0"); c.border = std_b
    c.alignment = Alignment(indent=1, vertical="center", wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 90


# ── Public entry point ────────────────────────────────────────────────────────

def build_workbook(df: pd.DataFrame, output_path: Path) -> None:
    """
    Build the full 3-sheet Excel workbook and write to output_path.

    Parameters
    ----------
    df          : enriched DataFrame from enrich.py
    output_path : destination .xlsx file path
    """
    wb = Workbook()
    as_of = datetime.now().strftime("%b %Y")

    # Sheet 1 — ASX Database
    ws_db = wb.active
    ws_db.title = "ASX Database"
    _build_database_sheet(ws_db, df, as_of)

    # Sheet 2 — Summary
    ws_sum = wb.create_sheet("Summary")
    _build_summary_sheet(ws_sum, df, as_of)

    # Sheet 3 — Legend
    ws_leg = wb.create_sheet("Legend")
    _build_legend_sheet(ws_leg)

    wb.active = wb["ASX Database"]
    wb.save(output_path)
